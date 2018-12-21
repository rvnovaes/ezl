from decimal import Decimal
from django.contrib.auth.models import User
from django.core.cache import cache
from django.utils import timezone
from enum import Enum
from celery import shared_task, current_task, chain
from openpyxl import load_workbook
from .models import ServicePriceTable, ImportServicePriceTable, PolicyPrice, CategoryPrice
from core.models import Office, State, City
from task.models import TypeTask
from lawsuit.models import CourtDistrict, CourtDistrictComplement
from .utils import remove_special_char

IMPORTED_IMPORT_SERVICE_PRICE_TABLE = 'imported_ispt_'
IMPORTED_WORKSHEET = 'imported_worksheet_'
PROCESS_PERCENT_IMPORT_SERVICE_PRICE_TABLE = 'process_percent_ispt_'
WORKSHEET_IN_PROCESS = 'worksheet_in_process_'
ERROR_PROCESS = 'ERROR_PROCESS_'


class ColumnIndex(Enum):
    correspondent = 0
    policy_price = 1
    type_task = 2
    client = 3
    state = 4
    court_district = 5
    court_district_complement = 6
    city = 7
    value = 8


def get_office_correspondent(row, xls_file):
    # escritorio correspondente
    office_correspondent = False

    if row[ColumnIndex.correspondent.value].value:
        office_name = remove_special_char(str(row[ColumnIndex.correspondent.value].value).strip())
        office_correspondent = Office.objects.filter(legal_name__unaccent__iexact=office_name).first()
        if not office_correspondent:                        
            xls_file.log = xls_file.log + ('Escritório correspondente %s não encontrado' % office_name) + ";"
            row[len(row) - 1]['errors'] = True
    return office_correspondent


def get_policy_price(row, xls_file, office_session):
    # Tipo de Preço
    policy_price = False
    if row[ColumnIndex.policy_price.value].value:
        name_policy_price = remove_special_char(str(row[ColumnIndex.policy_price.value].value).strip())
        policy_price = PolicyPrice.objects.filter(office=office_session,
                                                  name__unaccent__iexact=name_policy_price).first()
        if not policy_price:
            xls_file.log = xls_file.log + ('Tipo de preço %s não encontrado' % name_policy_price) + ";"
            row[len(row) - 1]['errors'] = True
    else:
        policy_price = PolicyPrice.objects.filter(office=office_session,
                                                  category=CategoryPrice.POSTPAID).first()
    return policy_price


def get_type_task(row, xls_file):
    # serviço
    type_task = False
    if row[ColumnIndex.type_task.value].value:
        name_service = remove_special_char(str(row[ColumnIndex.type_task.value].value).strip())
        type_task = TypeTask.objects.filter(name__unaccent__iexact=name_service).first()
        if not type_task:
            xls_file.log = xls_file.log + ('Tipo de serviço %s não encontrado' % name_service) + ";"
            row[len(row) - 1]['errors'] = True
    return type_task


def get_client(row, xls_file, office_session):
    # cliente
    client = False    
    if row[ColumnIndex.client.value].value:
        client_cleaned = remove_special_char(str(row[ColumnIndex.client.value].value).strip())
        client = office_session.persons.filter(legal_name__unaccent__iexact=client_cleaned, is_customer=True).first()
        if not client:
            xls_file.log = xls_file.log + ('Cliente %s não encontrado' % client_cleaned + ";")
            row[len(row) - 1]['errors'] = True
    return client


def get_office_network(row, xls_file, office_session):
    # rede utiliza a mesma coluna de correspondente, por isso pega o mesmo indice do enum
    office_network = False
    office_network_cleaned = remove_special_char(str(row[ColumnIndex.correspondent.value].value).strip())
    if row[ColumnIndex.correspondent.value].value:
        office_network = office_session.network_members.filter(name__unaccent__iexact=office_network_cleaned).first()
    if not office_network:
        xls_file.log = xls_file.log + ('Rede %s não encontrado' % office_network_cleaned + ";")
        row[len(row) - 1]['errors'] = True
    return office_network


def get_court_district(row, xls_file, state):
    # Comarca
    court_district = False
    if row[ColumnIndex.court_district.value].value and state:
        court_district_name = remove_special_char(str(row[ColumnIndex.court_district.value].value).strip())
        court_district = CourtDistrict.objects.filter(name__unaccent__iexact=court_district_name, state=state.pk).first()
        if not court_district:
            msg_error = '%s - %s' % (court_district_name, state.initials) if state else '%s' % court_district_name
            xls_file.log = xls_file.log + 'Comarca ' + msg_error + ' não encontrada' + ";"
            row[len(row) - 1]['errors'] = True
    return court_district


def get_court_district_complement(row, xls_file, court_district):
    # Complemento de Comarca
    court_district_complement = False
    if row[ColumnIndex.court_district_complement.value].value and court_district:
        complement_name = remove_special_char(str(row[ColumnIndex.court_district_complement.value].value).strip())
        court_district_complement = CourtDistrictComplement.objects.filter(
            name__unaccent__iexact=complement_name, court_district_id=court_district.pk).first()
        if not court_district_complement:
            msg_error = '%s - %s' % (complement_name, court_district.name) if court_district else '%s' % complement_name
            xls_file.log = xls_file.log + 'Complemento de comarca ' + msg_error + ' não encontrado' + ";"
            row[len(row) - 1]['errors'] = True
    return court_district_complement


def get_state(row, xls_file):
    # uf            
    state = False
    if row[ColumnIndex.state.value].value:
        uf = remove_special_char(str(row[ColumnIndex.state.value].value).strip())
        state = State.objects.filter(initials__iexact=uf).first()
        if not state:
            xls_file.log = xls_file.log + ('UF %s não encontrada' % uf) + ";"
            row[len(row) - 1]['errors'] = True
    return state


def get_city(row, xls_file, state):
    # uf
    city = False
    if row[ColumnIndex.city.value].value and state:
        city_name = remove_special_char(str(row[ColumnIndex.city.value].value).strip())
        city = City.objects.filter(name__unaccent__iexact=city_name, state=state).first()
        if not city:
            xls_file.log = xls_file.log + ('Cidade %s não encontrada para a UF %s' % (city_name, state.initials)) + ";"
            row[len(row) - 1]['errors'] = True
    return city


def get_service_value(row, xls_file, office_correspondent, policy_price, court_district, state, client):
    value = 0
    try:                        
        if type(row[ColumnIndex.value.value].value) == str:
            value = Decimal(row[ColumnIndex.value.value].value.replace('R$\xa0', '').replace(',', '.'))
        else:                        
            value = row[ColumnIndex.value.value].value
    except:        
        xls_file.log = xls_file.log + (
            'Valor {} inválido. Verificar escritório {}, comarca {}, estado {}, cliente {}.'.format(
                row[ColumnIndex.value.value].value, office_correspondent, policy_price, court_district, state,
                client)) + ";"
        row[len(row) - 1]['errors'] = True
    finally:
        return value


def row_is_valid(office_correspondent, policy_price, type_task, row_errors):
    return all([office_correspondent, policy_price, type_task, not row_errors])


def update_or_create_service_price_table(xls_file, office_session, user_session, office_correspondent, policy_price,
                                         office_network, type_task, court_district, court_district_complement, state,
                                         city, client, value):
    service_price_table = ServicePriceTable.objects.filter(
        office=office_session,
        office_correspondent=office_correspondent.pk,
        office_network=office_network,
        policy_price=policy_price,
        type_task=type_task.pk,
        court_district=court_district.pk if court_district else None,
        court_district_complement=court_district_complement.pk if court_district_complement else None,
        city=city.pk if city else None,
        client=client.pk if client else None,
        state=state.pk if state else None).first()
    if not service_price_table:                    
        service_price_table = ServicePriceTable.objects.create(
            office=office_session,
            office_correspondent=office_correspondent,
            office_network=office_network,
            policy_price=policy_price,
            type_task=type_task,
            court_district=court_district if court_district else None,
            court_district_complement=court_district_complement if court_district_complement else None,
            city=city if city else None,
            client=client if client else None,
            state=state if state else None,
            value=value,
            create_user=user_session
        )
    else:                        
        if value != service_price_table.value:
            xls_file.log = xls_file.log + ("Tabela de preço do escritório {}, serviço {} teve seu valor atualizado de "
                                           "{} para {}".format(office_correspondent, type_task,
                                                               service_price_table.value, value)) + ";"
            service_price_table.value = value
            service_price_table.save()


@shared_task(bind=True)
def import_xls_service_price_table(self, file_id):
    try:
        self.update_state(state="PROGRESS")
        xls_file = ImportServicePriceTable.objects.get(pk=file_id)
        xls_file.log = " "
        # chaves para acessar dados em cache
        imported_cache_key = IMPORTED_IMPORT_SERVICE_PRICE_TABLE + str(xls_file.pk)
        imported_worksheet_key = IMPORTED_WORKSHEET + str(xls_file.pk)
        percent_imported_cache_key = PROCESS_PERCENT_IMPORT_SERVICE_PRICE_TABLE + str(xls_file.pk)
        worksheet_in_process_key = WORKSHEET_IN_PROCESS + str(xls_file.pk)
        error_process_key = ERROR_PROCESS + str(xls_file.pk)        
        cache.set(imported_cache_key, False, timeout=None)
        cache.set(error_process_key, False, timeout=None)
        user_session = User.objects.get(pk=xls_file.create_user.pk)
        office_session = Office.objects.get(pk=xls_file.office.pk)       
        wb = load_workbook(xls_file.file_xls.file, data_only=True) 
        i = 0
        for sheet in wb.worksheets:
            total = sheet.max_row
            i = 0
            process_percent = 0
            cache.set(worksheet_in_process_key, sheet.title, timeout=None)
            cache.set(imported_worksheet_key, False, timeout=None)            
            for row in sheet.iter_rows(min_row=2):
                row += ({'errors': False}, )
                policy_price = get_policy_price(row, xls_file, office_session)
                if policy_price and policy_price.category in [CategoryPrice.NETWORK.name, CategoryPrice.PUBLIC.name]:
                    office_correspondent = office_session
                else:
                    office_correspondent = get_office_correspondent(row, xls_file)
                office_network = None
                if policy_price and policy_price.category == CategoryPrice.NETWORK.name:
                    office_network = get_office_network(row, xls_file, office_session)
                type_task = get_type_task(row, xls_file)
                state = get_state(row, xls_file)
                court_district = get_court_district(row, xls_file, state)
                court_district_complement = get_court_district_complement(row, xls_file, court_district)
                city = get_city(row, xls_file, state)
                client = get_client(row, xls_file, office_session)
                if row_is_valid(office_correspondent, policy_price, type_task, row[len(row) - 1]['errors']):
                    value = get_service_value(row, xls_file, office_correspondent, policy_price, court_district,
                                              state, client)
                    update_or_create_service_price_table(xls_file, office_session, user_session, office_correspondent,
                                                         policy_price, office_network, type_task, court_district,
                                                         court_district_complement, state, city, client, value)
                i = i + 1
                process_percent = int(100 * float(i) / float(total))
                cache.set(percent_imported_cache_key, process_percent, timeout=None)
            cache.set(imported_worksheet_key, True, timeout=None)                                
        cache.set(imported_cache_key, True, timeout=None)
    except FileNotFoundError as ex:
        cache.set(error_process_key, True, timeout=None)        
        xls_file.log = xls_file.log + " Planilha não encontrada Erro: " + str(ex.args[0]) + ";"        
    except Exception as ex:        
        cache.set(error_process_key, True, timeout=None)        
        xls_file.log = xls_file.log + " Planilha: " + sheet.title + " Linha " + str(i + 1) + "Erro: " + ex.args[0] + ";"        
    finally:        
        xls_file.end = timezone.now()
        xls_file.save()
        delete_imported_xls_service_price_table.apply_async(([xls_file.pk]), countdown=60)


@shared_task(bind=True)
def delete_imported_xls_service_price_table(self, xls_file_pk):
    ImportServicePriceTable.objects.filter(pk=xls_file_pk).delete()
